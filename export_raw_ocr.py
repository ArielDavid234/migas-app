"""
export_raw_ocr.py
-----------------
Llama al OCR y exporta los datos CRUDOS (sin parsear) a Excel:
  - Una hoja "Palabras" por foto  → cada token con Left (X), Top (Y) y Texto
  - Una hoja "Lineas"  por foto   → las líneas que el propio OCR detectó (agrupación nativa)
  - Una hoja "TextoRaw" por foto  → el texto plano devuelto por OCR

Uso: py export_raw_ocr.py
"""

import os
import sys

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

# Importamos las funciones internas del módulo (sin modificarlo)
from utils.ocr_scan import (
    _ocr_with_ocrspace_overlay,   # noqa: F401 — acceso directo al dato crudo
)
from config import OCR_SPACE_API_KEY

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── configuración ─────────────────────────────────────────────────────────────
PHOTOS  = ["2.jpeg", "3.jpeg", "4.jpeg"]
OUT_DIR = os.path.join(ROOT, "exports")
os.makedirs(OUT_DIR, exist_ok=True)

# ── estilos ───────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
TIT_FONT = Font(bold=True, size=11, color="1F4E79")
ALT_FILL = PatternFill("solid", fgColor="D9E1F2")
SIDE     = Side(style="thin", color="BFBFBF")
BORDER   = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)
CENTER   = Alignment(horizontal="center", vertical="center")
LEFT     = Alignment(horizontal="left",   vertical="center")


def _write_title(ws, text, n_cols):
    ws.append([text])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1).font      = TIT_FONT
    ws.cell(row=1, column=1).alignment = CENTER


def _write_header(ws, headers, row=3):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = CENTER
        c.border    = BORDER


def _write_row(ws, values, row_idx, excel_row):
    fill = ALT_FILL if row_idx % 2 == 0 else None
    for col, val in enumerate(values, 1):
        c = ws.cell(row=excel_row, column=col, value=val)
        c.border    = BORDER
        c.alignment = LEFT if col == len(values) else CENTER
        if fill:
            c.fill = fill


def _auto_col_widths(ws, headers, min_row=3):
    for col_idx, h in enumerate(headers, 1):
        max_len = len(h)
        for row in ws.iter_rows(min_row=min_row, max_col=col_idx, min_col=col_idx):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
api_key = OCR_SPACE_API_KEY.strip()
wb = openpyxl.Workbook()
wb.remove(wb.active)

for photo in PHOTOS:
    path = os.path.join(ROOT, photo)
    if not os.path.exists(path):
        print(f"  [OMITIDA] {photo} — no encontrada")
        continue

    print(f"  Escaneando {photo} ...", end=" ", flush=True)
    data = _ocr_with_ocrspace_overlay(path, api_key)
    words = data["overlay_words"]   # list of {text, left, top}
    lines = data["overlay_lines"]   # list of {top, text, words:[...]}
    raw   = data["raw_text"]

    slug = photo.replace(".", "_")          # "2_jpeg"
    n_words = len(words)
    n_lines = len(lines)
    print(f"OK — {n_words} palabras, {n_lines} líneas OCR")

    # ── Hoja 1: Palabras crudas ───────────────────────────────────────────────
    ws_w = wb.create_sheet(title=f"{slug}_palabras")
    HEADERS_W = ["#", "Top (Y)", "Left (X)", "Texto"]
    _write_title(ws_w, f"Foto: {photo}  |  Palabras crudas OCR  ({n_words} tokens)", len(HEADERS_W))
    ws_w.append([])   # fila 2 vacía
    _write_header(ws_w, HEADERS_W, row=3)

    # Ordenar por Top (Y) luego Left (X) — orden visual top-to-bottom, left-to-right
    sorted_words = sorted(words, key=lambda w: (w["top"], w["left"]))
    for i, w in enumerate(sorted_words, 1):
        _write_row(ws_w, [i, w["top"], w["left"], w["text"]], i, 3 + i)

    _auto_col_widths(ws_w, HEADERS_W)

    # ── Hoja 2: Líneas OCR ────────────────────────────────────────────────────
    ws_l = wb.create_sheet(title=f"{slug}_lineas")
    HEADERS_L = ["# Línea", "Top (Y)", "# Palabras", "Texto completo", "Palabras (X: texto) →"]
    _write_title(ws_l, f"Foto: {photo}  |  Líneas detectadas por OCR  ({n_lines} líneas)", len(HEADERS_L))
    ws_l.append([])
    _write_header(ws_l, HEADERS_L, row=3)

    for i, line in enumerate(lines, 1):
        words_detail = "  |  ".join(
            f"X={w['left']} : {w['text']}" for w in line["words"]
        )
        _write_row(ws_l,
                   [i, line["top"], len(line["words"]), line["text"], words_detail],
                   i, 3 + i)

    _auto_col_widths(ws_l, HEADERS_L)

    # ── Hoja 3: Texto plano ───────────────────────────────────────────────────
    ws_t = wb.create_sheet(title=f"{slug}_texto_raw")
    ws_t.column_dimensions["A"].width = 100
    ws_t.append([f"=== Texto plano OCR — {photo} ==="])
    ws_t.cell(row=1, column=1).font = TIT_FONT
    ws_t.append([])
    for line_num, line_text in enumerate(raw.splitlines(), 1):
        ws_t.append([line_text])

# ── guardar ───────────────────────────────────────────────────────────────────
out_path = os.path.join(OUT_DIR, "raw_ocr_data.xlsx")
wb.save(out_path)
print(f"\n  [XLSX]  {out_path}")
print("  Hojas por foto: _palabras  (tokens X/Y)  |  _lineas  (agrupación OCR)  |  _texto_raw")
