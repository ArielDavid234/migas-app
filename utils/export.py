"""Export helpers — generate Excel (.xlsx) and PDF reports."""

import os
from datetime import date, timedelta
from config import DATA_DIR
from database.db import get_session
from database.models import Sale, Expense, Product, Category, ShiftType, Schedule, User, UserRole
from sqlalchemy import func

EXPORT_DIR = os.path.join(DATA_DIR, "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


def _ts():
    from datetime import datetime
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ────────── Excel ──────────

def export_sales_excel(from_date: date, to_date: date) -> str:
    """Export sales for a date range to Excel. Returns file path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    session = get_session()
    try:
        sales = session.query(Sale).filter(
            Sale.date >= from_date, Sale.date <= to_date
        ).order_by(Sale.date, Sale.created_at).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"

        # Header style
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        headers = ["Fecha", "Turno", "Descripción", "Monto", "Cafetería"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for i, s in enumerate(sales, 2):
            ws.cell(row=i, column=1, value=s.date.strftime("%d/%m/%Y"))
            ws.cell(row=i, column=2, value="Mañana" if s.shift == ShiftType.MORNING else "Noche")
            ws.cell(row=i, column=3, value=s.description or "")
            ws.cell(row=i, column=4, value=s.amount).number_format = '#,##0.00'
            ws.cell(row=i, column=5, value="Sí" if s.is_cafeteria else "No")

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        # Total row
        total_row = len(sales) + 2
        ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=sum(s.amount for s in sales)).number_format = '#,##0.00'
        ws.cell(row=total_row, column=4).font = Font(bold=True)

        path = os.path.join(EXPORT_DIR, f"ventas_{_ts()}.xlsx")
        wb.save(path)
        return path
    finally:
        session.close()


def export_inventory_excel() -> str:
    """Export current inventory to Excel. Returns file path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    session = get_session()
    try:
        products = session.query(Product).join(Category).order_by(Category.name, Product.name).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        headers = ["Producto", "Categoría", "Stock", "Mín. Stock", "Precio", "Costo", "Vencimiento"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for i, p in enumerate(products, 2):
            ws.cell(row=i, column=1, value=p.name)
            ws.cell(row=i, column=2, value=p.category.name)
            ws.cell(row=i, column=3, value=p.stock)
            ws.cell(row=i, column=4, value=p.min_stock)
            ws.cell(row=i, column=5, value=p.price).number_format = '#,##0.00'
            ws.cell(row=i, column=6, value=p.cost).number_format = '#,##0.00'
            ws.cell(row=i, column=7, value=p.expiry_date.strftime("%d/%m/%Y") if p.expiry_date else "—")

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        path = os.path.join(EXPORT_DIR, f"inventario_{_ts()}.xlsx")
        wb.save(path)
        return path
    finally:
        session.close()


def export_expenses_excel(from_date: date, to_date: date) -> str:
    """Export expenses for a date range to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    session = get_session()
    try:
        expenses = session.query(Expense).filter(
            Expense.date >= from_date, Expense.date <= to_date
        ).order_by(Expense.date, Expense.created_at).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos"

        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        headers = ["Fecha", "Descripción", "Monto", "Mercancía"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for i, ex in enumerate(expenses, 2):
            ws.cell(row=i, column=1, value=ex.date.strftime("%d/%m/%Y"))
            ws.cell(row=i, column=2, value=ex.description)
            ws.cell(row=i, column=3, value=ex.amount).number_format = '#,##0.00'
            ws.cell(row=i, column=4, value="Sí" if ex.is_merchandise else "No")

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        total_row = len(expenses) + 2
        ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=sum(e.amount for e in expenses)).number_format = '#,##0.00'
        ws.cell(row=total_row, column=3).font = Font(bold=True)

        path = os.path.join(EXPORT_DIR, f"gastos_{_ts()}.xlsx")
        wb.save(path)
        return path
    finally:
        session.close()


# ────────── PDF ──────────

def export_daily_summary_excel(for_date: date) -> str:
    """Generate an Excel workbook with the daily sales/expenses summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    session = get_session()
    try:
        sales = session.query(Sale).filter(Sale.date == for_date).order_by(Sale.created_at).all()
        expenses = session.query(Expense).filter(Expense.date == for_date).order_by(Expense.created_at).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Diario"

        white_bold = Font(bold=True, color="FFFFFF", size=11)
        blue_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        red_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
        green_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        center = Alignment(horizontal="center")

        # Title
        ws.append([f"Resumen Diario — {for_date.strftime('%d/%m/%Y')}"])
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.append([])

        # ── Ventas section ──
        ws.append(["VENTAS"])
        ws.cell(ws.max_row, 1).font = white_bold
        ws.cell(ws.max_row, 1).fill = blue_fill
        hdr_row = ws.max_row + 1
        for col, h in enumerate(["Turno", "Descripción", "Monto", "Cafetería"], 1):
            c = ws.cell(hdr_row, col, h)
            c.font = white_bold
            c.fill = blue_fill
            c.alignment = center
        for s in sales:
            ws.append([
                "Mañana" if s.shift == ShiftType.MORNING else "Noche",
                s.description or "—",
                s.amount,
                "Sí" if s.is_cafeteria else "No",
            ])
            ws.cell(ws.max_row, 3).number_format = "#,##0.00"
        total_s = sum(s.amount for s in sales)
        ws.append(["", "TOTAL", total_s, ""])
        ws.cell(ws.max_row, 2).font = Font(bold=True)
        ws.cell(ws.max_row, 3).font = Font(bold=True)
        ws.cell(ws.max_row, 3).number_format = "#,##0.00"
        ws.append([])

        # ── Gastos section ──
        ws.append(["GASTOS"])
        ws.cell(ws.max_row, 1).font = white_bold
        ws.cell(ws.max_row, 1).fill = red_fill
        hdr_row2 = ws.max_row + 1
        for col, h in enumerate(["Descripción", "Monto", "Mercancía"], 1):
            c = ws.cell(hdr_row2, col, h)
            c.font = white_bold
            c.fill = red_fill
            c.alignment = center
        for ex in expenses:
            ws.append([ex.description, ex.amount, "Sí" if ex.is_merchandise else "No"])
            ws.cell(ws.max_row, 2).number_format = "#,##0.00"
        total_e = sum(e.amount for e in expenses)
        ws.append(["TOTAL", total_e, ""])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.cell(ws.max_row, 2).font = Font(bold=True)
        ws.cell(ws.max_row, 2).number_format = "#,##0.00"
        ws.append([])

        # ── Balance ──
        balance = total_s - total_e
        sign = "+" if balance >= 0 else ""
        ws.append([f"BALANCE DEL DÍA: {sign}${balance:,.2f}"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
        ws.cell(ws.max_row, 1).fill = green_fill if balance >= 0 else red_fill
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12, color="FFFFFF")

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        path = os.path.join(EXPORT_DIR, f"resumen_{for_date.strftime('%Y%m%d')}_{_ts()}.xlsx")
        wb.save(path)
        return path
    finally:
        session.close()


def export_shift_summary_excel(
    for_date: date,
    shift_type: ShiftType,
    user_name: str,
    start_time: str,
    end_time: str,
) -> str:
    """Generate an Excel workbook with the balance for a specific worker shift."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    session = get_session()
    try:
        shift_label = "Mañana" if shift_type == ShiftType.MORNING else "Noche"
        sales = (
            session.query(Sale)
            .filter(Sale.date == for_date, Sale.shift == shift_type)
            .order_by(Sale.created_at)
            .all()
        )
        expenses = (
            session.query(Expense)
            .filter(Expense.date == for_date)
            .order_by(Expense.created_at)
            .all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Turno"

        white_bold = Font(bold=True, color="FFFFFF", size=11)
        blue_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        red_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
        purple_fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
        green_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        center = Alignment(horizontal="center")

        # Info header
        ws.append([f"Reporte de Turno — {shift_label} ({start_time}–{end_time})"])
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.append([f"Trabajador: {user_name}   |   Fecha: {for_date.strftime('%d/%m/%Y')}"])
        ws.cell(2, 1).font = Font(italic=True, size=11)
        ws.append([])

        # ── Ventas del turno ──
        ws.append(["VENTAS DEL TURNO"])
        ws.cell(ws.max_row, 1).font = white_bold
        ws.cell(ws.max_row, 1).fill = blue_fill
        hdr_row = ws.max_row + 1
        for col, h in enumerate(["Descripción", "Monto", "Cafetería"], 1):
            c = ws.cell(hdr_row, col, h)
            c.font = white_bold
            c.fill = blue_fill
            c.alignment = center
        for s in sales:
            ws.append([s.description or "—", s.amount, "Sí" if s.is_cafeteria else "No"])
            ws.cell(ws.max_row, 2).number_format = "#,##0.00"
        if not sales:
            ws.append(["Sin ventas registradas", 0, ""])
        total_s = sum(s.amount for s in sales)
        ws.append(["TOTAL", total_s, ""])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.cell(ws.max_row, 2).font = Font(bold=True)
        ws.cell(ws.max_row, 2).number_format = "#,##0.00"
        ws.append([])

        # ── Gastos del día (referencia) ──
        ws.append(["GASTOS DEL DÍA (REFERENCIA)"])
        ws.cell(ws.max_row, 1).font = white_bold
        ws.cell(ws.max_row, 1).fill = red_fill
        hdr_row2 = ws.max_row + 1
        for col, h in enumerate(["Descripción", "Monto", "Mercancía"], 1):
            c = ws.cell(hdr_row2, col, h)
            c.font = white_bold
            c.fill = red_fill
            c.alignment = center
        for ex in expenses:
            ws.append([ex.description, ex.amount, "Sí" if ex.is_merchandise else "No"])
            ws.cell(ws.max_row, 2).number_format = "#,##0.00"
        if not expenses:
            ws.append(["Sin gastos registrados", 0, ""])
        total_e = sum(e.amount for e in expenses)
        ws.append(["TOTAL", total_e, ""])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.cell(ws.max_row, 2).font = Font(bold=True)
        ws.cell(ws.max_row, 2).number_format = "#,##0.00"
        ws.append([])

        # ── Balance ──
        balance = total_s - total_e
        sign = "+" if balance >= 0 else ""
        ws.append([f"BALANCE DEL TURNO: {sign}${balance:,.2f}"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(ws.max_row, 1).fill = green_fill if balance >= 0 else red_fill

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        path = os.path.join(
            EXPORT_DIR,
            f"turno_{for_date.strftime('%Y%m%d')}_{shift_label.lower()}_{_ts()}.xlsx",
        )
        wb.save(path)
        return path
    finally:
        session.close()


def export_schedule_excel(week_start: date, filter_user_id: int | None = None) -> str:
    """
    Generate an Excel workbook with the weekly schedule table.
    If filter_user_id is given, only that worker's row is included.
    Returns the file path.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    DAYS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    week_end = week_start + timedelta(days=6)

    session = get_session()
    try:
        q = session.query(User).filter_by(is_active=True, role=UserRole.WORKER).order_by(User.name)
        if filter_user_id:
            q = q.filter(User.id == filter_user_id)
        workers = q.all()

        schedules = session.query(Schedule).filter(Schedule.week_start == week_start).all()
        sched_map: dict[int, dict[int, Schedule]] = {}
        for s in schedules:
            if filter_user_id and s.user_id != filter_user_id:
                continue
            sched_map.setdefault(s.user_id, {})[s.day_of_week] = s

        from config import GAS_STATION_NAME
        from datetime import datetime as _dt

        wb = Workbook()
        ws = wb.active
        ws.title = "Horario Semanal"

        white_bold = Font(bold=True, color="FFFFFF", size=11)
        blue_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        name_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Title rows
        ws.append([GAS_STATION_NAME])
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.append([f"Horario Semanal: {week_start.strftime('%d/%m/%Y')} \u2014 {week_end.strftime('%d/%m/%Y')}"])
        ws.cell(2, 1).font = Font(italic=True, size=11)
        ws.append([])

        # Header row
        hdr = ["Trabajador"] + [
            f"{DAYS_ES[i]}\n{(week_start + timedelta(days=i)).strftime('%d/%m')}"
            for i in range(7)
        ]
        ws.append(hdr)
        hdr_row_idx = ws.max_row
        for col in range(1, 9):
            c = ws.cell(hdr_row_idx, col)
            c.font = white_bold
            c.fill = blue_fill
            c.alignment = center
        ws.row_dimensions[hdr_row_idx].height = 30

        # Data rows
        for row_num, w in enumerate(workers):
            user_sched = sched_map.get(w.id, {})
            row = [w.name]
            for day_idx in range(7):
                s = user_sched.get(day_idx)
                row.append(f"{s.start_time} \u2013 {s.end_time}" if s else "Libre")
            ws.append(row)
            data_row_idx = ws.max_row
            # Name cell style
            ws.cell(data_row_idx, 1).font = Font(bold=True)
            ws.cell(data_row_idx, 1).fill = name_fill
            # Alternate row shading
            for col in range(2, 9):
                c = ws.cell(data_row_idx, col)
                c.alignment = center
                if row_num % 2 == 1:
                    c.fill = alt_fill
            ws.row_dimensions[data_row_idx].height = 20

        if not workers:
            ws.append(["No hay trabajadores con horario asignado esta semana."])

        # Column widths
        ws.column_dimensions["A"].width = 22
        for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col_letter].width = 16

        ws.append([])
        ws.append([f"Generado el {_dt.now().strftime('%d/%m/%Y %H:%M')}"])
        ws.cell(ws.max_row, 1).font = Font(italic=True, size=9)

        path = os.path.join(EXPORT_DIR, f"horario_{week_start.strftime('%Y%m%d')}_{_ts()}.xlsx")
        wb.save(path)
        return path
    finally:
        session.close()


def export_report_excel_bytes(reports_data: list, date_val, shift_label: str) -> str:
    """
    Generate an Excel workbook from already-loaded report dicts and save to disk.
    Each dict must have keys: report, user_name, cigarettes, lottery, checks, tips, specials.
    Returns the file path.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    white_bold = Font(bold=True, color="FFFFFF", size=11)
    blue_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    green_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    purple_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
    teal_fill = PatternFill(start_color="004D40", end_color="004D40", fill_type="solid")
    orange_fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    total_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="BDBDBD")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, "strftime") else str(date_val)

    def _hdr(ws, col, row, value, fill):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = fill
        c.font = white_bold
        c.alignment = center
        c.border = thin_border
        return c

    def _cell(ws, col, row, value, bold=False, number_format=None, fill=None, align=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold)
        c.border = thin_border
        if number_format:
            c.number_format = number_format
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        return c

    def _auto_width(ws, min_w=10, max_w=40):
        from openpyxl.cell import MergedCell
        for col in ws.columns:
            first = next((c for c in col if not isinstance(c, MergedCell)), None)
            if first is None:
                continue
            max_len = max((len(str(c.value or "")) for c in col if not isinstance(c, MergedCell)), default=0)
            ws.column_dimensions[first.column_letter].width = max(min_w, min(max_len + 4, max_w))

    def _worker_header(ws, user_name, shift_name, fill):
        ws.append(["  Trabajador: " + user_name + "   |   Turno: " + shift_name])
        r = ws.max_row
        ws.cell(r, 1).fill = fill
        ws.cell(r, 1).font = Font(bold=True, size=10, color="FFFFFF")
        ws.row_dimensions[r].height = 16

    # Sheet 1 - Resumen
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.append(["Reporte - " + shift_label + "   |   " + date_str])
    ws1.cell(1, 1).font = Font(bold=True, size=14, color="0D47A1")
    ws1.merge_cells("A1:E1")
    ws1.row_dimensions[1].height = 22
    ws1.append([])

    for rd in reports_data:
        shift_name = "Manana" if rd["report"].shift.name == "MORNING" else "Noche"
        checks_total = sum(c.amount for c in rd["checks"])
        tips_total = sum(t.amount for t in rd["tips"])
        lottery_total = sum(l.amount + l.lotto_amount for l in rd["lottery"])
        cig_sold = sum(c.sold for c in rd["cigarettes"])
        _worker_header(ws1, rd["user_name"], shift_name, blue_fill)
        hdr_row = ws1.max_row + 1
        _hdr(ws1, 1, hdr_row, "Concepto", blue_fill)
        _hdr(ws1, 2, hdr_row, "Valor", blue_fill)
        items = [
            ("Total Loteria (Scratch + Lotto)", "$" + format(lottery_total, ",.2f")),
            ("Total Cheques", "$" + format(checks_total, ",.2f")),
            ("Total Propinas", "$" + format(tips_total, ",.2f")),
            ("Cigarros vendidos (unidades)", str(cig_sold)),
        ]
        for i, (concept, val) in enumerate(items):
            r = ws1.max_row + 1
            _cell(ws1, 1, r, concept, fill=alt_fill if i % 2 else None)
            _cell(ws1, 2, r, val, fill=alt_fill if i % 2 else None, align=center)
        ws1.append([])
    _auto_width(ws1)

    # Sheet 2 - Cigarros
    ws2 = wb.create_sheet("Cigarros")
    ws2.append(["Contabilidad de Cigarros - " + shift_label + "   |   " + date_str])
    ws2.cell(1, 1).font = Font(bold=True, size=13, color="0D47A1")
    ws2.merge_cells("A1:D1")
    ws2.row_dimensions[1].height = 20
    ws2.append([])
    for rd in reports_data:
        shift_name = "Manana" if rd["report"].shift.name == "MORNING" else "Noche"
        _worker_header(ws2, rd["user_name"], shift_name, green_fill)
        if not rd["cigarettes"]:
            ws2.append(["  Sin datos"])
            ws2.append([])
            continue
        hdr_row = ws2.max_row + 1
        for col, h in enumerate(["Marca", "Inicio", "Vendidos", "Resto"], 1):
            _hdr(ws2, col, hdr_row, h, green_fill)
        total_sold = 0
        for i, c in enumerate(rd["cigarettes"]):
            r = ws2.max_row + 1
            _cell(ws2, 1, r, c.brand, fill=alt_fill if i % 2 else None)
            _cell(ws2, 2, r, c.boxes_start, fill=alt_fill if i % 2 else None, align=center)
            _cell(ws2, 3, r, c.sold, fill=alt_fill if i % 2 else None, align=center)
            _cell(ws2, 4, r, c.boxes_end, fill=alt_fill if i % 2 else None, align=center)
            total_sold += c.sold
        r = ws2.max_row + 1
        _cell(ws2, 1, r, "TOTAL VENDIDOS", bold=True, fill=total_fill)
        _cell(ws2, 3, r, total_sold, bold=True, fill=total_fill, align=center)
        ws2.append([])
    _auto_width(ws2)

    # Sheet 3 - Loteria
    ws3 = wb.create_sheet("Loteria")
    ws3.append(["Ventas de Loteria - " + shift_label + "   |   " + date_str])
    ws3.cell(1, 1).font = Font(bold=True, size=13, color="0D47A1")
    ws3.merge_cells("A1:D1")
    ws3.row_dimensions[1].height = 20
    ws3.append([])
    for rd in reports_data:
        shift_name = "Manana" if rd["report"].shift.name == "MORNING" else "Noche"
        _worker_header(ws3, rd["user_name"], shift_name, purple_fill)
        if not rd["lottery"]:
            ws3.append(["  Sin datos"])
            ws3.append([])
            continue
        hdr_row = ws3.max_row + 1
        for col, h in enumerate(["Scratch", "Ventas Scratch", "Lotto", "Total"], 1):
            _hdr(ws3, col, hdr_row, h, purple_fill)
        grand_total = 0.0
        for i, l in enumerate(rd["lottery"]):
            total = l.amount + l.lotto_amount
            r = ws3.max_row + 1
            _cell(ws3, 1, r, l.scratch_name, fill=alt_fill if i % 2 else None)
            _cell(ws3, 2, r, l.amount, number_format="#,##0.00", fill=alt_fill if i % 2 else None, align=center)
            _cell(ws3, 3, r, l.lotto_amount, number_format="#,##0.00", fill=alt_fill if i % 2 else None, align=center)
            _cell(ws3, 4, r, total, number_format="#,##0.00", fill=alt_fill if i % 2 else None, align=center)
            grand_total += total
        r = ws3.max_row + 1
        _cell(ws3, 1, r, "TOTAL", bold=True, fill=total_fill)
        _cell(ws3, 4, r, grand_total, bold=True, number_format="#,##0.00", fill=total_fill, align=center)
        ws3.append([])
    _auto_width(ws3)

    # Sheet 4 - Cheques
    ws4 = wb.create_sheet("Cheques")
    ws4.append(["Cheques Emitidos - " + shift_label + "   |   " + date_str])
    ws4.cell(1, 1).font = Font(bold=True, size=13, color="0D47A1")
    ws4.merge_cells("A1:B1")
    ws4.row_dimensions[1].height = 20
    ws4.append([])
    for rd in reports_data:
        shift_name = "Manana" if rd["report"].shift.name == "MORNING" else "Noche"
        _worker_header(ws4, rd["user_name"], shift_name, teal_fill)
        if not rd["checks"]:
            ws4.append(["  Sin cheques"])
            ws4.append([])
            continue
        hdr_row = ws4.max_row + 1
        _hdr(ws4, 1, hdr_row, "Descripcion", teal_fill)
        _hdr(ws4, 2, hdr_row, "Monto", teal_fill)
        total = 0.0
        for i, ch in enumerate(rd["checks"]):
            r = ws4.max_row + 1
            _cell(ws4, 1, r, ch.description or "-", fill=alt_fill if i % 2 else None)
            _cell(ws4, 2, r, ch.amount, number_format="#,##0.00", fill=alt_fill if i % 2 else None, align=center)
            total += ch.amount
        r = ws4.max_row + 1
        _cell(ws4, 1, r, "TOTAL", bold=True, fill=total_fill)
        _cell(ws4, 2, r, total, bold=True, number_format="#,##0.00", fill=total_fill, align=center)
        ws4.append([])
    _auto_width(ws4)

    # Sheet 5 - Propinas
    ws5 = wb.create_sheet("Propinas")
    ws5.append(["Propinas - " + shift_label + "   |   " + date_str])
    ws5.cell(1, 1).font = Font(bold=True, size=13, color="0D47A1")
    ws5.merge_cells("A1:C1")
    ws5.row_dimensions[1].height = 20
    ws5.append([])
    hdr_row = ws5.max_row + 1
    for col, h in enumerate(["Trabajador", "Turno", "Total Propinas"], 1):
        _hdr(ws5, col, hdr_row, h, orange_fill)
    grand_tips = 0.0
    for i, rd in enumerate(reports_data):
        shift_name = "Manana" if rd["report"].shift.name == "MORNING" else "Noche"
        tips_total = sum(t.amount for t in rd["tips"])
        r = ws5.max_row + 1
        _cell(ws5, 1, r, rd["user_name"], fill=alt_fill if i % 2 else None)
        _cell(ws5, 2, r, shift_name, fill=alt_fill if i % 2 else None, align=center)
        _cell(ws5, 3, r, tips_total, number_format="#,##0.00", fill=alt_fill if i % 2 else None, align=center)
        grand_tips += tips_total
    r = ws5.max_row + 1
    _cell(ws5, 1, r, "TOTAL", bold=True, fill=total_fill)
    _cell(ws5, 3, r, grand_tips, bold=True, number_format="#,##0.00", fill=total_fill, align=center)
    _auto_width(ws5)

    # Sheet 6 - Especiales
    ws6 = wb.create_sheet("Especiales")
    ws6.append(["Productos Especiales - " + shift_label + "   |   " + date_str])
    ws6.cell(1, 1).font = Font(bold=True, size=13, color="0D47A1")
    ws6.merge_cells("A1:C1")
    ws6.row_dimensions[1].height = 20
    ws6.append([])
    for rd in reports_data:
        shift_name = "Manana" if rd["report"].shift.name == "MORNING" else "Noche"
        _worker_header(ws6, rd["user_name"], shift_name, teal_fill)
        if not rd["specials"]:
            ws6.append(["  Sin datos"])
            ws6.append([])
            continue
        hdr_row = ws6.max_row + 1
        for col, h in enumerate(["Producto", "Vendidos", "Quedan"], 1):
            _hdr(ws6, col, hdr_row, h, teal_fill)
        for i, s in enumerate(rd["specials"]):
            r = ws6.max_row + 1
            _cell(ws6, 1, r, s.item_name, fill=alt_fill if i % 2 else None)
            _cell(ws6, 2, r, s.sold, fill=alt_fill if i % 2 else None, align=center)
            _cell(ws6, 3, r, s.remaining, fill=alt_fill if i % 2 else None, align=center)
        ws6.append([])
    _auto_width(ws6)

    date_str_file = date_val.strftime("%Y%m%d") if hasattr(date_val, "strftime") else str(date_val)
    safe_label = shift_label.lower().replace(" ", "_").replace("/", "-")
    path = os.path.join(EXPORT_DIR, f"reporte_{date_str_file}_{safe_label}_{_ts()}.xlsx")
    wb.save(path)
    return path


# ────────── Importación masiva de inventario ──────────

def export_inventory_import_template() -> str:
    """
    Generate a blank Excel template the user can fill to bulk-import products.
    Returns the file path.
    Columns: Nombre*, Categoría*, Stock, Stock Mínimo, Precio Venta, Costo,
             Consignación (Sí/No), Fecha Vencimiento (DD/MM/AAAA)
    A second sheet lists all available categories.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    session = get_session()
    try:
        categories = session.query(Category).order_by(Category.name).all()
        cat_names = [c.name for c in categories]
    finally:
        session.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    req_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")  # darker = required
    center = Alignment(horizontal="center", vertical="center")

    COLUMNS = [
        ("Nombre *",               20),
        ("Categoría *",            22),
        ("Stock",                  10),
        ("Stock Mínimo",           14),
        ("Precio Venta",           14),
        ("Costo",                  12),
        ("Consignación (Sí/No)",   20),
        ("Fecha Vencimiento",      20),
    ]

    for col_idx, (title, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col_idx, value=title)
        c.font = hdr_font
        c.fill = req_fill if title.endswith("*") else hdr_fill
        c.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Example row
    ws.append(["Coca-Cola 330ml", cat_names[0] if cat_names else "Bebidas", 24, 5, 1.50, 0.80, "No", ""])
    example_font = Font(italic=True, color="757575")
    for col_idx in range(1, 9):
        ws.cell(row=2, column=col_idx).font = example_font

    # Note row
    ws.append([])
    ws.append(["* Campos obligatorios.  Fecha formato DD/MM/AAAA (ej: 31/12/2026).  "
               "Consignación: Sí o No.  Columnas vacías usan valores por defecto (Stock=0, Mín=2, Precio=0, Costo=0)."])
    note_row = ws.max_row
    ws.cell(note_row, 1).font = Font(italic=True, size=9, color="616161")
    ws.merge_cells(f"A{note_row}:H{note_row}")

    # Sheet 2: categories reference
    ws2 = wb.create_sheet("Categorías Disponibles")
    ws2.append(["Categorías disponibles (copia el nombre exacto en la columna 'Categoría')"])
    ws2.cell(1, 1).font = Font(bold=True, size=12, color="0D47A1")
    ws2.column_dimensions["A"].width = 40
    for name in cat_names:
        ws2.append([name])

    path = os.path.join(EXPORT_DIR, f"plantilla_inventario_{_ts()}.xlsx")
    wb.save(path)
    return path


def import_inventory_from_excel(filepath: str) -> dict:
    """
    Read a filled inventory template and bulk-create/update products.
    Returns a dict: {created: int, updated: int, skipped: int, errors: list[str]}
    Rules:
      - Nombre + Categoría required.
      - If product with same name (case-insensitive) already exists → update stock (add delta).
      - Category must match an existing category (case-insensitive).
    """
    from openpyxl import load_workbook
    from datetime import datetime as _dt

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    session = get_session()
    try:
        # Build category lookup (lowercase → Category object)
        cat_map = {c.name.lower(): c for c in session.query(Category).all()}

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row_num, row in enumerate(rows, start=2):
            # Skip blank or comment rows
            if not row or not row[0]:
                skipped += 1
                continue

            raw_name = str(row[0]).strip()
            raw_cat  = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            # Skip the example row
            if raw_name.startswith("Coca-Cola") and row_num == 2:
                skipped += 1
                continue

            if not raw_name:
                skipped += 1
                continue
            if not raw_cat:
                errors.append(f"Fila {row_num}: '{raw_name}' — falta categoría.")
                continue

            cat = cat_map.get(raw_cat.lower())
            if cat is None:
                errors.append(f"Fila {row_num}: '{raw_name}' — categoría '{raw_cat}' no encontrada.")
                continue

            def _to_int(val, default=0):
                try:
                    return int(float(val)) if val not in (None, "", " ") else default
                except (ValueError, TypeError):
                    return default

            def _to_float(val, default=0.0):
                try:
                    return float(val) if val not in (None, "", " ") else default
                except (ValueError, TypeError):
                    return default

            stock     = _to_int(row[2] if len(row) > 2 else None, 0)
            min_stock = _to_int(row[3] if len(row) > 3 else None, 2)
            price     = _to_float(row[4] if len(row) > 4 else None, 0.0)
            cost      = _to_float(row[5] if len(row) > 5 else None, 0.0)
            consign_raw = str(row[6]).strip().lower() if len(row) > 6 and row[6] else "no"
            is_consignment = consign_raw in ("sí", "si", "s", "yes", "y", "true", "1")

            expiry_date = None
            if len(row) > 7 and row[7]:
                raw_exp = str(row[7]).strip()
                for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
                    try:
                        expiry_date = _dt.strptime(raw_exp, fmt).date()
                        break
                    except ValueError:
                        pass

            # Check if product already exists (same name, case-insensitive)
            existing = session.query(Product).filter(
                Product.name.ilike(raw_name),
                Product.category_id == cat.id,
            ).first()

            if existing:
                existing.stock += stock
                existing.min_stock = min_stock if min_stock != 2 else existing.min_stock
                existing.price = price if price > 0 else existing.price
                existing.cost = cost if cost > 0 else existing.cost
                if expiry_date:
                    existing.expiry_date = expiry_date
                updated += 1
            else:
                new_prod = Product(
                    name=raw_name,
                    category_id=cat.id,
                    stock=stock,
                    min_stock=min_stock,
                    price=price,
                    cost=cost,
                    is_consignment=is_consignment,
                    expiry_date=expiry_date,
                )
                session.add(new_prod)
                created += 1

        session.commit()
    except Exception as exc:
        session.rollback()
        errors.append(f"Error crítico al guardar: {exc}")
    finally:
        session.close()
        try:
            wb.close()
        except Exception:
            pass

    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


# ────────── Escaneo de reportes (descuento de stock) ──────────

def export_scan_report_template() -> str:
    """
    Generate a blank Excel template for the scan-report feature.
    Columns: Nombre del Producto*, Cantidad a descontar*
    Returns the file path.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    COLUMNS = [("Nombre del Producto *", 30), ("Cantidad a Descontar *", 24)]
    for col_idx, (title, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col_idx, value=title)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # Example rows
    example_font = Font(italic=True, color="757575")
    examples = [("Coca-Cola 330ml", 5), ("Agua 500ml", 12), ("Aceite Motor 4L", 2)]
    for row_idx, (nm, qty) in enumerate(examples, 2):
        ws.cell(row_idx, 1, nm).font = example_font
        ws.cell(row_idx, 2, qty).font = example_font

    note_row = len(examples) + 3
    ws.cell(note_row, 1,
            "* El nombre debe coincidir exactamente con el registrado en el inventario (sin importar mayúsculas).")
    ws.cell(note_row, 1).font = Font(italic=True, size=9, color="616161")
    ws.merge_cells(f"A{note_row}:B{note_row}")

    path = os.path.join(EXPORT_DIR, f"plantilla_reporte_{_ts()}.xlsx")
    wb.save(path)
    return path


def parse_scan_report(filepath: str) -> dict:
    """
    Read a scan-report Excel and return preview data.
    Returns:
      {
        "rows": [{"product_id", "name", "qty_remove", "current_stock", "remaining",
                  "error": str|None}],
        "parse_errors": [str],
      }
    Each row has an "error" field if the product was not found or qty is invalid.
    Does NOT modify the DB — caller applies changes after user confirmation.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    parse_errors: list[str] = []

    session = get_session()
    try:
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            raw_name = str(row[0]).strip()
            if not raw_name:
                continue

            # Quantity
            try:
                qty = int(float(str(row[1]).strip())) if len(row) > 1 and row[1] not in (None, "") else 0
            except (ValueError, TypeError):
                parse_errors.append(f"Fila {row_num}: cantidad inválida «{row[1]}»")
                continue

            if qty <= 0:
                parse_errors.append(f"Fila {row_num}: la cantidad debe ser mayor a 0")
                continue

            prod = session.query(Product).filter(
                Product.name.ilike(raw_name),
                Product.status.in_(["active", "ACTIVE"]),
            ).first()

            if prod is None:
                rows.append({
                    "product_id": None,
                    "name": raw_name,
                    "qty_remove": qty,
                    "current_stock": None,
                    "remaining": None,
                    "error": "Producto no encontrado en el inventario",
                })
            else:
                remaining = prod.stock - qty
                rows.append({
                    "product_id": prod.id,
                    "name": prod.name,
                    "qty_remove": qty,
                    "current_stock": prod.stock,
                    "remaining": remaining,
                    "error": "Stock insuficiente" if remaining < 0 else None,
                })
    finally:
        session.close()
        try:
            wb.close()
        except Exception:
            pass

    return {"rows": rows, "parse_errors": parse_errors}


def apply_scan_report(rows: list, applied_by_id) -> dict:
    """
    Apply confirmed scan-report rows: deduct stock from products.
    Skips rows with errors. Returns {applied: int, skipped: int}.
    """
    applied = 0
    skipped = 0
    session = get_session()
    try:
        for r in rows:
            if r.get("error") or r["product_id"] is None:
                skipped += 1
                continue
            prod = session.query(Product).get(r["product_id"])
            if prod:
                prod.stock = max(0, prod.stock - r["qty_remove"])
                applied += 1
        session.commit()
    finally:
        session.close()
    return {"applied": applied, "skipped": skipped}


def apply_department_scan_report(rows: list, report_date, user_id) -> dict:
    """
    Save confirmed department-sale rows (from DEPARTMENT REPORT OCR scan) to the DB.
    Creates a DepartmentSaleReport header + one DepartmentSaleRow per item.
    Returns {saved: int, report_id: int}.
    """
    from database.models import DepartmentSaleReport, DepartmentSaleRow
    from datetime import date as _date

    session = get_session()
    try:
        report = DepartmentSaleReport(
            user_id=user_id,
            report_date=report_date if report_date else _date.today(),
        )
        session.add(report)
        session.flush()  # get report.id before adding rows

        for r in rows:
            try:
                items_val = int(r.get("items") or 0)
            except (ValueError, TypeError):
                items_val = 0
            session.add(DepartmentSaleRow(
                dept_report_id=report.id,
                dept_num=str(r.get("dept_num", "") or ""),
                description=str(r.get("description", "") or ""),
                items=items_val,
                sales_gross=float(r.get("sales_gross") or 0.0),
                refunds=float(r.get("refunds") or 0.0),
                discounts=float(r.get("discounts") or 0.0),
                net_sales=float(r.get("net_sales") or 0.0),
            ))

        session.commit()
        return {"saved": len(rows), "report_id": report.id}
    finally:
        session.close()
